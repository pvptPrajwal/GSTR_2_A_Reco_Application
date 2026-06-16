"""
================================================================================
  GSTR-2A Reconciliation Tool  —  Desktop GUI App  v2.0
  Changes in v2.0:
    ✅ Select MULTIPLE GST Portal files at once
    ✅ Select MULTIPLE Purchase Book files at once
    ✅ All files combined — flat single reconciliation (no month split)
    ✅ Vendor Name from both GST Portal and Purchase Book
    ✅ Cross_Match sheet — GSTINs in both missing sheets
    ✅ Vendor_Summary sheet — per-vendor PB vs GST comparison
    ✅ Output filename includes today date (no overwrite)
    ✅ Incomplete_Records sheet — PB rows missing GSTIN or Invoice Number
================================================================================
"""

import re, sys, os, threading, datetime
from pathlib import Path
from collections import defaultdict
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

OUTPUT_FILE  = "Final_Reco_Report.xlsx"
KEY_COLS     = ["GSTIN", "Invoice Number"]
VALUE_COLS   = ["Taxable Value", "IGST", "CGST", "SGST"]
MISMATCH_TOL = 1.0
GSTIN_RE     = re.compile(r"^\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z\d]$")
MONTHS       = ["January","February","March","April","May","June",
                "July","August","September","October","November","December"]
MONTH_ABBR   = ["jan","feb","mar","apr","may","jun",
                "jul","aug","sep","oct","nov","dec"]

COL_ALIASES = {
    "gstin of supplier":"GSTIN","supplier gstin":"GSTIN","gstin of sez supplier":"GSTIN",
    "party gstin":"GSTIN","vendor gstin":"GSTIN","gstin":"GSTIN",
    "invoice no":"Invoice Number","invoice no.":"Invoice Number",
    "invoice number":"Invoice Number","note no":"Invoice Number","note no.":"Invoice Number",
    "bill of entry no":"Invoice Number","bill of entry no.":"Invoice Number",
    "bill no":"Invoice Number","bill number":"Invoice Number",
    "document no":"Invoice Number","doc no":"Invoice Number",
    "invoice date":"Invoice Date","note date":"Invoice Date",
    "bill of entry date":"Invoice Date","date":"Invoice Date",
    "taxable value":"Taxable Value","taxable amount":"Taxable Value",
    "assessable value":"Taxable Value",
    "igst":"IGST","igst paid":"IGST","igst amount":"IGST",
    "integrated tax":"IGST","integrated tax paid":"IGST",
    "cgst":"CGST","cgst paid":"CGST","cgst amount":"CGST",
    "central tax":"CGST","central tax paid":"CGST",
    "sgst":"SGST","sgst paid":"SGST","sgst amount":"SGST","sgst/ut":"SGST",
    "sgst/ut paid":"SGST","state tax":"SGST","state/ut tax paid":"SGST",
    "eligibility for itc":"ITC Eligibility","itc availability":"ITC Eligibility",
    "itc eligible":"ITC Eligibility","itc":"ITC Eligibility",
    "trade/legal name":"Vendor Name","trade name":"Vendor Name",
    "legal name":"Vendor Name","name of the supplier":"Vendor Name",
    "supplier name":"Vendor Name",
    "party name":"Vendor Name","vendor name":"Vendor Name",
    "supplier":"Vendor Name","party":"Vendor Name",
    "creditor name":"Vendor Name",
}

# ─────────────────────────────────────────────────────────────────────────────
# COLOUR PALETTE
# ─────────────────────────────────────────────────────────────────────────────

C = {
    "bg":"#F0F4F8","card":"#FFFFFF","primary":"#1F4E79","primary_lt":"#2E75B6",
    "success":"#1F5C2E","warning":"#C55A11","danger":"#C00000","purple":"#7030A0",
    "text":"#1A1A2E","subtext":"#6B7280","border":"#D1D5DB",
    "green_bg":"#D9EAD3","orange_bg":"#FCE5CD","red_bg":"#FFE0E0","purple_bg":"#EAD1DC",
    "white":"#FFFFFF",
}

SHEET_COLOURS = {
    "Matched":"1F5C2E","Mismatched":"C55A11",
    "Missing_in_GST":"C00000","Missing_in_PB":"7030A0",
    "Summary":"1F4E79","Duplicates_Removed":"595959",
    "Incomplete_Records":"7F6000",
}

THIN = Border(
    left=Side(style="thin",color="B8CCE4"), right=Side(style="thin",color="B8CCE4"),
    top=Side(style="thin",color="B8CCE4"),  bottom=Side(style="thin",color="B8CCE4"),
)
RED_FILL    = PatternFill("solid",start_color="FFD7D7",end_color="FFD7D7")
YELLOW_FILL = PatternFill("solid",start_color="FFF2CC",end_color="FFF2CC")
ORANGE_FILL = PatternFill("solid",start_color="FFE0B2",end_color="FFE0B2")
AMBER_FILL  = PatternFill("solid",start_color="FFF2CC",end_color="FFF2CC")

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 1 — MONTH DETECTION FROM FILENAME
# ─────────────────────────────────────────────────────────────────────────────

def detect_month_from_filename(filepath: str) -> str | None:
    name = Path(filepath).stem.lower()
    for i, (full, abbr) in enumerate(zip([m.lower() for m in MONTHS], MONTH_ABBR)):
        if full in name or abbr in name:
            return MONTHS[i]
    nums = re.findall(r'\b(0?[1-9]|1[0-2])\b', name)
    if nums:
        return MONTHS[int(nums[0]) - 1]
    return None


def detect_month_from_data(df: pd.DataFrame) -> str | None:
    if "Invoice Date" not in df.columns:
        return None
    dates = pd.to_datetime(df["Invoice Date"], errors="coerce", dayfirst=True)
    dates = dates.dropna()
    if dates.empty:
        return None
    most_common_month = dates.dt.month.mode()[0]
    return MONTHS[most_common_month - 1]


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 2 — CORE DATA HELPERS
# ─────────────────────────────────────────────────────────────────────────────

_STRIP_RE = re.compile(r"[^A-Za-z0-9]")

def clean_invoice(val):
    if pd.isna(val) or val is None: return ""
    return _STRIP_RE.sub("", str(val).strip()).upper()

def clean_gstin(val):
    if pd.isna(val) or val is None: return ""
    return str(val).strip().upper()

def _get_col(df, col, default=0.0):
    if col is None or col not in df.columns:
        return pd.Series([default]*len(df), index=df.index)
    return pd.to_numeric(df[col], errors="coerce").fillna(0.0)

def _get_str_col(df, col, default=""):
    if col is None or col not in df.columns:
        return pd.Series([default]*len(df), index=df.index)
    return df[col].fillna(default).astype(str)

def _rename_cols(df):
    rename_map = {col: COL_ALIASES[str(col).strip().lower()]
                  for col in df.columns
                  if str(col).strip().lower() in COL_ALIASES}
    return df.rename(columns=rename_map)

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 3 — FILE LOADING
# ─────────────────────────────────────────────────────────────────────────────

def _find_header_row(filepath, sheet_name, scan_rows=15):
    raw = pd.read_excel(filepath, sheet_name=sheet_name,
                        header=None, nrows=scan_rows, dtype=str)
    invoice_keys = {"invoice no","invoice no.","invoice number","note no",
                    "note no.","bill of entry no","bill no","bill number"}
    value_keys   = {"taxable value","taxable amount","igst","igst paid",
                    "cgst","cgst paid","sgst","sgst paid"}
    best_row, best_score = 0, 0
    for idx, row in raw.iterrows():
        cells = {str(v).strip().lower() for v in row if pd.notna(v) and str(v).strip()}
        if bool(cells & invoice_keys) and bool(cells & value_keys):
            return int(idx)
        score = sum(1 for c in cells if c in COL_ALIASES)
        if score > best_score:
            best_score = score; best_row = int(idx)
    return best_row


def _extract_sheet(filepath, sheet_name):
    hdr = _find_header_row(filepath, sheet_name)
    raw = pd.read_excel(filepath, sheet_name=sheet_name, header=hdr, dtype=str)
    raw.dropna(how="all", inplace=True)
    raw.reset_index(drop=True, inplace=True)
    if raw.empty: return None
    raw = _rename_cols(raw)
    if "Invoice Number" not in raw.columns: return None
    if not any(c in raw.columns for c in VALUE_COLS): return None

    chunk = pd.DataFrame()
    chunk["GSTIN"]           = raw["GSTIN"].apply(clean_gstin) if "GSTIN" in raw.columns else pd.Series([""]*len(raw))
    chunk["Vendor Name"]     = _get_str_col(raw, "Vendor Name")
    chunk["Invoice Number"]  = raw["Invoice Number"].apply(clean_invoice)
    chunk["Invoice Date"]    = _get_str_col(raw, "Invoice Date")
    chunk["Taxable Value"]   = _get_col(raw, "Taxable Value")
    chunk["IGST"]            = _get_col(raw, "IGST")
    chunk["CGST"]            = _get_col(raw, "CGST")
    chunk["SGST"]            = _get_col(raw, "SGST")
    chunk["ITC Eligibility"] = _get_str_col(raw, "ITC Eligibility", "N/A")
    chunk = chunk[chunk["Invoice Number"] != ""]
    return chunk if not chunk.empty else None


def load_gst_file(filepath: str, log) -> pd.DataFrame:
    xl = pd.ExcelFile(filepath)
    frames = []
    for sh in xl.sheet_names:
        chunk = _extract_sheet(filepath, sh)
        if chunk is not None:
            frames.append(chunk)
            log(f"    '{sh}' → {len(chunk):,} rows")
        else:
            log(f"    '{sh}' → skipped")
    if not frames:
        raise ValueError(f"No usable data in: {Path(filepath).name}")
    df = pd.concat(frames, ignore_index=True)
    log(f"    Total: {len(df):,} rows from {Path(filepath).name}")
    return df


def load_pb_file(filepath: str, log) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load Purchase Book file.
    Returns (clean_df, incomplete_df)
      clean_df      — rows that have BOTH GSTIN and Invoice Number
      incomplete_df — rows missing GSTIN or Invoice Number (new!)
    """
    xl    = pd.ExcelFile(filepath)
    sheet = xl.sheet_names[0]

    # Read raw WITHOUT cleaning so we can capture incomplete rows
    raw = pd.read_excel(filepath, sheet_name=sheet, dtype=str)
    raw.dropna(how="all", inplace=True)
    raw.reset_index(drop=True, inplace=True)
    raw = _rename_cols(raw)

    for col in VALUE_COLS:
        if col not in raw.columns: raw[col] = 0.0
    if "GSTIN"        not in raw.columns: raw["GSTIN"]        = ""
    if "Invoice Date" not in raw.columns: raw["Invoice Date"] = ""
    if "Vendor Name"  not in raw.columns: raw["Vendor Name"]  = ""
    if "Invoice Number" not in raw.columns: raw["Invoice Number"] = ""

    # Clean the key columns
    raw["GSTIN"]          = raw["GSTIN"].apply(clean_gstin)
    raw["Invoice Number"] = raw["Invoice Number"].apply(clean_invoice)
    raw["Invoice Date"]   = raw["Invoice Date"].fillna("").astype(str)
    raw["Vendor Name"]    = raw["Vendor Name"].fillna("").astype(str)
    for col in VALUE_COLS:
        raw[col] = pd.to_numeric(raw[col], errors="coerce").fillna(0.0)

    # ── Split into complete and incomplete ────────────────────────────────────
    missing_gstin   = raw["GSTIN"] == ""
    missing_invoice = raw["Invoice Number"] == ""
    incomplete_mask = missing_gstin | missing_invoice

    incomplete_df = raw[incomplete_mask].copy()
    clean_df      = raw[~incomplete_mask].copy()

    # Add a reason column to incomplete records
    def _reason(row):
        reasons = []
        if row["GSTIN"] == "":          reasons.append("Missing GSTIN")
        if row["Invoice Number"] == "": reasons.append("Missing Invoice Number")
        return " | ".join(reasons)
    if not incomplete_df.empty:
        incomplete_df["Reason"] = incomplete_df.apply(_reason, axis=1)

    keep = KEY_COLS + ["Vendor Name", "Invoice Date"] + VALUE_COLS
    clean_df = clean_df[[c for c in keep if c in clean_df.columns]].copy()

    log(f"    {len(clean_df):,} complete rows from {Path(filepath).name}")
    if len(incomplete_df) > 0:
        log(f"    ⚠️  {len(incomplete_df):,} incomplete row(s) — missing GSTIN or Invoice Number")

    return clean_df.reset_index(drop=True), incomplete_df.reset_index(drop=True)


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 4 — GROUP FILES BY MONTH (kept for future use)
# ─────────────────────────────────────────────────────────────────────────────

def group_files_by_month(filepaths: list[str], file_type: str, log) -> dict[str, list[str]]:
    groups = defaultdict(list)
    for fp in filepaths:
        month = detect_month_from_filename(fp)
        if month:
            log(f"  [{file_type}] {Path(fp).name} → {month}")
        else:
            log(f"  [{file_type}] {Path(fp).name} → month not detected, will check data")
            month = "__check_data__"
        groups[month].append(fp)
    return dict(groups)


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 5 — DUPLICATE DETECTION  (temporarily disabled)
# ─────────────────────────────────────────────────────────────────────────────

# def detect_and_remove_duplicates(df: pd.DataFrame, label: str, log):
#     check_cols  = KEY_COLS + VALUE_COLS
#     is_dup      = df.duplicated(subset=check_cols, keep="first")
#     dup_df      = df[is_dup].copy()
#     clean_df    = df[~is_dup].copy()
#     if len(dup_df) > 0:
#         log(f"  ⚠️  [{label}] {len(dup_df)} duplicate row(s) removed")
#     else:
#         log(f"  [{label}] No duplicates ✅")
#     return clean_df.reset_index(drop=True), dup_df.reset_index(drop=True)


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 6 — AGGREGATION
# ─────────────────────────────────────────────────────────────────────────────

def aggregate(df: pd.DataFrame, label: str, log) -> pd.DataFrame:
    agg = df.groupby(KEY_COLS, as_index=False, sort=False)[VALUE_COLS].sum().round(2)
    if "Vendor Name" in df.columns:
        vn  = df.groupby(KEY_COLS, sort=False)["Vendor Name"].first().reset_index()
        agg = agg.merge(vn, on=KEY_COLS, how="left")
    if "ITC Eligibility" in df.columns:
        priority = {"Ineligible":0,"No":1,"Yes":2,"N/A":3}
        worst    = lambda v: min([str(x).strip() for x in v], key=lambda x: priority.get(x,99))
        itc_agg  = df.groupby(KEY_COLS, sort=False)["ITC Eligibility"].apply(worst).reset_index()
        agg      = agg.merge(itc_agg, on=KEY_COLS, how="left")
    log(f"  [{label}] {len(df):,} rows → {len(agg):,} unique invoices")
    return agg.reset_index(drop=True)


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 7 — GSTIN VALIDATION
# ─────────────────────────────────────────────────────────────────────────────

def validate_gstins(df: pd.DataFrame, label: str, log) -> list[str]:
    warnings = []
    invalid  = df[(df["GSTIN"] != "") &
                  (~df["GSTIN"].apply(lambda g: bool(GSTIN_RE.match(g))))]["GSTIN"].unique()
    for g in invalid:
        msg = f"  [{label}] Invalid GSTIN: {g}"
        warnings.append(msg); log(msg)
    return warnings


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 8 — RECONCILIATION ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def reconcile(pb: pd.DataFrame, gst: pd.DataFrame, log):
    pb_m  = pb.copy()
    gst_m = gst.copy()
    if "Vendor Name" in pb_m.columns:
        pb_m.rename(columns={"Vendor Name":"Vendor Name (PB)"}, inplace=True)
    if "Vendor Name" in gst_m.columns:
        gst_m.rename(columns={"Vendor Name":"Vendor Name (GST)"}, inplace=True)

    merged = pd.merge(pb_m, gst_m, on=KEY_COLS, how="outer",
                      suffixes=("_PB","_GST"), indicator=True)

    if "Vendor Name (PB)"  not in merged.columns: merged["Vendor Name (PB)"]  = ""
    if "Vendor Name (GST)" not in merged.columns: merged["Vendor Name (GST)"] = ""
    merged["Vendor Name (PB)"]  = merged["Vendor Name (PB)"].fillna("")
    merged["Vendor Name (GST)"] = merged["Vendor Name (GST)"].fillna("")

    both = merged[merged["_merge"]=="both"].copy()
    for col in VALUE_COLS:
        both[f"{col} Diff"] = (both[f"{col}_PB"] - both[f"{col}_GST"]).round(2)
    both["Total Tax Diff"] = both[[f"{c} Diff" for c in ["IGST","CGST","SGST"]]].abs().sum(axis=1).round(2)

    matched    = both[both["Total Tax Diff"] <= MISMATCH_TOL].copy()
    mismatched = both[both["Total Tax Diff"] >  MISMATCH_TOL].copy()

    matched.drop(columns=[f"{c} Diff" for c in VALUE_COLS]+["Total Tax Diff","_merge"],
                 errors="ignore", inplace=True)

    # Enforce correct column order for Matched sheet
    matched_col_order = (
        KEY_COLS + ["Vendor Name (PB)"] +
        [f"{c}_PB" for c in VALUE_COLS] +
        ["Vendor Name (GST)"] +
        [f"{c}_GST" for c in VALUE_COLS] +
        ["ITC Eligibility"]
    )
    matched = matched[[c for c in matched_col_order if c in matched.columns]]

    mismatched.drop(columns=["_merge"], errors="ignore", inplace=True)

    # Enforce correct column order for Mismatched sheet
    mismatch_col_order = (
        KEY_COLS + ["Vendor Name (PB)"] +
        [f"{c}_PB" for c in VALUE_COLS] +
        ["Vendor Name (GST)"] +
        [f"{c}_GST" for c in VALUE_COLS] +
        [f"{c} Diff" for c in VALUE_COLS] +
        ["Total Tax Diff"] + ["ITC Eligibility"]
    )
    mismatched = mismatched[[c for c in mismatch_col_order if c in mismatched.columns]]

    miss_gst = merged[merged["_merge"]=="left_only"].copy()
    miss_gst.drop(columns="_merge", inplace=True)
    pb_cols  = KEY_COLS + ["Vendor Name (PB)"] + [f"{c}_PB" for c in VALUE_COLS]
    miss_gst = miss_gst[[c for c in pb_cols if c in miss_gst.columns]]
    miss_gst.columns = [c.replace("_PB","") for c in miss_gst.columns]

    miss_pb  = merged[merged["_merge"]=="right_only"].copy()
    miss_pb.drop(columns="_merge", inplace=True)
    gst_cols = KEY_COLS + ["Vendor Name (GST)"] + [f"{c}_GST" for c in VALUE_COLS]
    if "ITC Eligibility" in miss_pb.columns: gst_cols += ["ITC Eligibility"]
    miss_pb  = miss_pb[[c for c in gst_cols if c in miss_pb.columns]]
    miss_pb.columns = [c.replace("_GST","") for c in miss_pb.columns]

    log(f"  ✅ Matched:{len(matched):,}  🔴 Mismatch:{len(mismatched):,}  "
        f"⚠️ MissGST:{len(miss_gst):,}  ⚠️ MissPB:{len(miss_pb):,}")
    return matched, mismatched, miss_gst, miss_pb


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 9 — EXCEL FORMATTING HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _fmt_ws(ws, hex_col):
    hdr_fill = PatternFill("solid",start_color=hex_col,  end_color=hex_col)
    alt_fill = PatternFill("solid",start_color="EBF3FB", end_color="EBF3FB")
    wht_fill = PatternFill("solid",start_color="FFFFFF", end_color="FFFFFF")
    money    = {i for i,c in enumerate(ws[1],1) if c.value in VALUE_COLS}
    for r,row in enumerate(ws.iter_rows(),1):
        for ci,cell in enumerate(row,1):
            cell.border = THIN
            if r == 1:
                cell.fill=hdr_fill
                cell.font=Font(name="Arial",bold=True,color="FFFFFF",size=10)
                cell.alignment=Alignment(horizontal="center",vertical="center")
            else:
                rgb = getattr(getattr(cell.fill,"start_color",None),"rgb","")
                if rgb in ("00000000","FFFFFFFF",""):
                    cell.fill = alt_fill if r%2==0 else wht_fill
                cell.font=Font(name="Arial",size=10)
                cell.alignment=Alignment(
                    horizontal="right" if ci in money else "left",
                    vertical="center")
                if ci in money: cell.number_format="#,##0.00"
    for col_cells in ws.columns:
        w = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(w+4,45)
    ws.freeze_panes = "A2"


def _highlight_mismatch(ws, df):
    col_map = {cell.value:idx for idx,cell in enumerate(ws[1],1)}
    for r,(_,row) in enumerate(df.iterrows(),2):
        for col in VALUE_COLS:
            pb,gs = f"{col}_PB",f"{col}_GST"
            if pb in col_map and gs in col_map:
                if abs(float(row.get(pb,0) or 0)-float(row.get(gs,0) or 0))>0.01:
                    ws.cell(r,col_map[pb]).fill=RED_FILL
                    ws.cell(r,col_map[gs]).fill=RED_FILL


def _totals_row(ws):
    last=ws.max_row; tr=last+1
    ws.cell(tr,1).value="TOTAL"
    ws.cell(tr,1).font=Font(name="Arial",bold=True,size=10)
    ws.cell(tr,1).fill=YELLOW_FILL
    for ci,cell in enumerate(ws[1],1):
        if cell.value and any(v in str(cell.value) for v in VALUE_COLS+["Diff"]):
            cl=get_column_letter(ci); tc=ws.cell(tr,ci)
            tc.value=f"=SUM({cl}2:{cl}{last})"
            tc.font=Font(name="Arial",bold=True,size=10)
            tc.number_format="#,##0.00"
            tc.fill=YELLOW_FILL
            tc.alignment=Alignment(horizontal="right")


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 9B — INCOMPLETE RECORDS SHEET WRITER
# ─────────────────────────────────────────────────────────────────────────────

def _write_incomplete_sheet(ws, df: pd.DataFrame):
    """
    Write the Incomplete_Records sheet.
    These are Purchase Book rows that are missing GSTIN or Invoice Number.
    They cannot be reconciled and need to be fixed in the source file.
    """
    hdr_fill = PatternFill("solid", start_color="7F6000", end_color="7F6000")
    alt_fill = PatternFill("solid", start_color="FFFBEB", end_color="FFFBEB")
    wht_fill = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    red_fill = PatternFill("solid", start_color="FFD7D7", end_color="FFD7D7")

    # Title
    ws["A1"] = "Incomplete Records — Purchase Book rows missing GSTIN or Invoice Number"
    ws["A1"].font = Font(name="Arial", bold=True, size=12, color="7F6000")
    ws["A2"] = "These rows could NOT be reconciled. Please fix them in your Purchase Book file."
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="C00000")
    ws.row_dimensions[1].height = 22

    if df.empty:
        ws["A4"] = "✅  No incomplete records found — all rows have GSTIN and Invoice Number."
        ws["A4"].font = Font(name="Arial", size=11, color="1F5C2E", bold=True)
        return

    # Decide columns to show — keep all original cols + Reason at end
    show_cols = [c for c in df.columns if c != "Reason"] + ["Reason"]
    show_cols = [c for c in show_cols if c in df.columns]

    # Header row at row 4
    hdr_row = 4
    for ci, h in enumerate(show_cols, 1):
        cell = ws.cell(hdr_row, ci, h)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN

    # Data rows
    for r, (_, row) in enumerate(df.iterrows(), hdr_row + 1):
        for ci, h in enumerate(show_cols, 1):
            val  = row.get(h, "")
            cell = ws.cell(r, ci, val if pd.notna(val) else "")
            cell.border = THIN
            cell.font   = Font(name="Arial", size=10)
            # Highlight missing key cells in red
            if h == "GSTIN" and str(val).strip() == "":
                cell.fill = red_fill
            elif h == "Invoice Number" and str(val).strip() == "":
                cell.fill = red_fill
            elif h == "Reason":
                cell.fill = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
                cell.font = Font(name="Arial", size=10, bold=True, color="7F6000")
            else:
                cell.fill = alt_fill if r % 2 == 0 else wht_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto column widths
    for col_cells in ws.columns:
        w = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(w + 4, 40)

    ws.freeze_panes = "A5"


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 9C — VENDOR-WISE SUMMARY BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_vendor_summary(pb_agg: pd.DataFrame, gst_agg: pd.DataFrame) -> pd.DataFrame:
    pb_grp = pb_agg.groupby("GSTIN", as_index=False).agg(
        **{"Vendor Name (PB)" : ("Vendor Name", "first") if "Vendor Name" in pb_agg.columns else ("GSTIN","first"),
           "TV_PB"            : ("Taxable Value","sum"),
           "IGST_PB"          : ("IGST","sum"),
           "CGST_PB"          : ("CGST","sum"),
           "SGST_PB"          : ("SGST","sum"),}
    ).round(2)
    if "Vendor Name" not in pb_agg.columns:
        pb_grp["Vendor Name (PB)"] = ""

    gst_grp = gst_agg.groupby("GSTIN", as_index=False).agg(
        **{"Vendor Name (GST)" : ("Vendor Name","first") if "Vendor Name" in gst_agg.columns else ("GSTIN","first"),
           "TV_GST"            : ("Taxable Value","sum"),
           "IGST_GST"          : ("IGST","sum"),
           "CGST_GST"          : ("CGST","sum"),
           "SGST_GST"          : ("SGST","sum"),}
    ).round(2)
    if "Vendor Name" not in gst_agg.columns:
        gst_grp["Vendor Name (GST)"] = ""

    merged = pd.merge(pb_grp, gst_grp, on="GSTIN", how="outer")
    for col in ["Vendor Name (PB)","Vendor Name (GST)",
                "TV_PB","IGST_PB","CGST_PB","SGST_PB",
                "TV_GST","IGST_GST","CGST_GST","SGST_GST"]:
        if col not in merged.columns:
            merged[col] = "" if "Name" in col else 0.0
        if "Name" not in col:
            merged[col] = pd.to_numeric(merged[col], errors="coerce").fillna(0.0)

    merged["Diff_TV"]   = (merged["TV_PB"]   - merged["TV_GST"]).round(2)
    merged["Diff_IGST"] = (merged["IGST_PB"] - merged["IGST_GST"]).round(2)
    merged["Diff_CGST"] = (merged["CGST_PB"] - merged["CGST_GST"]).round(2)
    merged["Diff_SGST"] = (merged["SGST_PB"] - merged["SGST_GST"]).round(2)

    col_order = ["GSTIN","Vendor Name (PB)","Vendor Name (GST)",
                 "TV_PB","IGST_PB","CGST_PB","SGST_PB",
                 "TV_GST","IGST_GST","CGST_GST","SGST_GST",
                 "Diff_TV","Diff_IGST","Diff_CGST","Diff_SGST"]
    merged = merged[[c for c in col_order if c in merged.columns]]
    return merged.reset_index(drop=True)


def _write_vendor_summary_sheet(ws, vendor_df: pd.DataFrame, month: str):
    ws["B2"]      = f"Vendor-wise Summary — {month}"
    ws["B2"].font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    ws.row_dimensions[2].height = 24

    headers  = list(vendor_df.columns)
    pb_cols  = {"TV_PB","IGST_PB","CGST_PB","SGST_PB","Vendor Name (PB)"}
    gst_cols = {"TV_GST","IGST_GST","CGST_GST","SGST_GST","Vendor Name (GST)"}
    diff_cols= {"Diff_TV","Diff_IGST","Diff_CGST","Diff_SGST"}
    fills    = {
        "key" : PatternFill("solid",start_color="1F4E79",end_color="1F4E79"),
        "pb"  : PatternFill("solid",start_color="1F5C2E",end_color="1F5C2E"),
        "gst" : PatternFill("solid",start_color="7030A0",end_color="7030A0"),
        "diff": PatternFill("solid",start_color="C55A11",end_color="C55A11"),
    }
    for c, h in enumerate(headers, 2):
        cell = ws.cell(4, c, h)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN
        if h in pb_cols:     cell.fill = fills["pb"]
        elif h in gst_cols:  cell.fill = fills["gst"]
        elif h in diff_cols: cell.fill = fills["diff"]
        else:                cell.fill = fills["key"]

    money_cols = {i+2 for i, h in enumerate(headers)
                  if h not in {"GSTIN","Vendor Name (PB)","Vendor Name (GST)"}}
    alt_fill = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
    wht_fill = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    red_fill = PatternFill("solid", start_color="FFD7D7", end_color="FFD7D7")

    for r, (_, row) in enumerate(vendor_df.iterrows(), 5):
        for c, h in enumerate(headers, 2):
            val  = row[h]
            cell = ws.cell(r, c, val if pd.notna(val) else "")
            cell.border = THIN
            cell.font   = Font(name="Arial", size=10)
            cell.fill   = alt_fill if r % 2 == 0 else wht_fill
            if c in money_cols:
                cell.number_format = "#,##0.00"
                cell.alignment     = Alignment(horizontal="right", vertical="center")
                if h in diff_cols and isinstance(val, (int, float)) and abs(val) > 0.01:
                    cell.fill = red_fill
                    cell.font = Font(name="Arial", size=10, bold=True, color="C00000")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    last = 4 + len(vendor_df); tr = last + 1
    ws.cell(tr, 2).value = "TOTAL"
    ws.cell(tr, 2).font  = Font(name="Arial", bold=True, size=10)
    ws.cell(tr, 2).fill  = YELLOW_FILL
    for c, h in enumerate(headers, 2):
        if c in money_cols:
            cl = get_column_letter(c); tc = ws.cell(tr, c)
            tc.value = f"=SUM({cl}5:{cl}{last})"
            tc.font=Font(name="Arial",bold=True,size=10)
            tc.number_format="#,##0.00"; tc.fill=YELLOW_FILL
            tc.alignment=Alignment(horizontal="right")

    for col_cells in ws.columns:
        w = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(w+4, 30)
    ws.freeze_panes = "B5"


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 9D — CROSS MATCH
# ─────────────────────────────────────────────────────────────────────────────

def build_cross_match(miss_gst: pd.DataFrame, miss_pb: pd.DataFrame) -> pd.DataFrame:
    if miss_gst.empty or miss_pb.empty:
        return pd.DataFrame()
    gstins_in_miss_gst = set(miss_gst["GSTIN"].unique())
    gstins_in_miss_pb  = set(miss_pb["GSTIN"].unique())
    common_gstins      = gstins_in_miss_gst & gstins_in_miss_pb
    if not common_gstins:
        return pd.DataFrame()

    mg = miss_gst[miss_gst["GSTIN"].isin(common_gstins)].copy()
    mp = miss_pb [miss_pb ["GSTIN"].isin(common_gstins)].copy()

    pb_grp = mg.groupby("GSTIN", as_index=False)[VALUE_COLS].sum().round(2)
    pb_grp.rename(columns={"Taxable Value":"TV_PB","IGST":"IGST_PB","CGST":"CGST_PB","SGST":"SGST_PB"}, inplace=True)
    gst_grp = mp.groupby("GSTIN", as_index=False)[VALUE_COLS].sum().round(2)
    gst_grp.rename(columns={"Taxable Value":"TV_GST","IGST":"IGST_GST","CGST":"CGST_GST","SGST":"SGST_GST"}, inplace=True)

    def _first_name(df, col):
        if col not in df.columns:
            return pd.Series([""] * len(df["GSTIN"].unique()), index=df.groupby("GSTIN").first().index)
        return df.groupby("GSTIN")[col].first()

    vn_pb  = _first_name(mg, "Vendor Name (PB)") if "Vendor Name (PB)"  in mg.columns else _first_name(mg, "Vendor Name")
    vn_gst = _first_name(mp, "Vendor Name (GST)") if "Vendor Name (GST)" in mp.columns else _first_name(mp, "Vendor Name")

    vn_df = pd.DataFrame({
        "GSTIN"            : list(common_gstins),
        "Vendor Name (PB)" : [vn_pb.get(g, "")  for g in common_gstins],
        "Vendor Name (GST)": [vn_gst.get(g, "") for g in common_gstins],
    })
    result = vn_df.merge(pb_grp, on="GSTIN", how="left").merge(gst_grp, on="GSTIN", how="left")
    for col in ["TV_PB","IGST_PB","CGST_PB","SGST_PB","TV_GST","IGST_GST","CGST_GST","SGST_GST"]:
        result[col] = result[col].fillna(0.0)
    result["Diff_TV"]   = (result["TV_PB"]   - result["TV_GST"]).round(2)
    result["Diff_IGST"] = (result["IGST_PB"] - result["IGST_GST"]).round(2)
    result["Diff_CGST"] = (result["CGST_PB"] - result["CGST_GST"]).round(2)
    result["Diff_SGST"] = (result["SGST_PB"] - result["SGST_GST"]).round(2)
    result.sort_values("GSTIN", inplace=True)
    return result.reset_index(drop=True)


def _write_cross_match_sheet(ws, df: pd.DataFrame):
    ws["B2"]      = "Cross Match — GSTINs appearing in both Missing sheets"
    ws["B2"].font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    ws.row_dimensions[2].height = 24
    ws["B3"]      = "These vendors have some invoices missing on BOTH sides — needs attention"
    ws["B3"].font = Font(name="Arial", italic=True, size=9, color="C00000")

    if df.empty:
        ws["B5"]      = "✅  No common GSTINs found in both missing sheets."
        ws["B5"].font = Font(name="Arial", size=11, color="1F5C2E", bold=True)
        return

    pb_set   = {"TV_PB","IGST_PB","CGST_PB","SGST_PB","Vendor Name (PB)"}
    gst_set  = {"TV_GST","IGST_GST","CGST_GST","SGST_GST","Vendor Name (GST)"}
    diff_set = {"Diff_TV","Diff_IGST","Diff_CGST","Diff_SGST"}
    hfills   = {
        "pb"  : PatternFill("solid",start_color="C00000",end_color="C00000"),
        "gst" : PatternFill("solid",start_color="7030A0",end_color="7030A0"),
        "diff": PatternFill("solid",start_color="C55A11",end_color="C55A11"),
        "key" : PatternFill("solid",start_color="1F4E79",end_color="1F4E79"),
    }
    headers = list(df.columns); hdr_row = 5; money_ci = set()
    for c, h in enumerate(headers, 2):
        cell = ws.cell(hdr_row, c, h)
        cell.font=Font(name="Arial",bold=True,color="FFFFFF",size=10)
        cell.alignment=Alignment(horizontal="center",vertical="center")
        cell.border=THIN
        if h in pb_set:    cell.fill=hfills["pb"]
        elif h in gst_set: cell.fill=hfills["gst"]
        elif h in diff_set:cell.fill=hfills["diff"]
        else:              cell.fill=hfills["key"]
        if h not in {"GSTIN","Vendor Name (PB)","Vendor Name (GST)"}: money_ci.add(c)

    alt_fill=PatternFill("solid",start_color="EBF3FB",end_color="EBF3FB")
    wht_fill=PatternFill("solid",start_color="FFFFFF",end_color="FFFFFF")
    red_fill=PatternFill("solid",start_color="FFD7D7",end_color="FFD7D7")

    for r,(_, row) in enumerate(df.iterrows(), hdr_row+1):
        for c, h in enumerate(headers, 2):
            val=row[h]; cell=ws.cell(r,c,val if pd.notna(val) else "")
            cell.border=THIN; cell.font=Font(name="Arial",size=10)
            cell.fill=alt_fill if r%2==0 else wht_fill
            if c in money_ci:
                cell.number_format="#,##0.00"
                cell.alignment=Alignment(horizontal="right",vertical="center")
                if h in diff_set and isinstance(val,(int,float)) and abs(val)>0.01:
                    cell.fill=red_fill; cell.font=Font(name="Arial",size=10,bold=True,color="C00000")
            else:
                cell.alignment=Alignment(horizontal="left",vertical="center")

    last=hdr_row+len(df); tr=last+1
    ws.cell(tr,2).value="TOTAL"; ws.cell(tr,2).font=Font(name="Arial",bold=True,size=10)
    ws.cell(tr,2).fill=YELLOW_FILL
    for c,h in enumerate(headers,2):
        if c in money_ci:
            cl=get_column_letter(c); tc=ws.cell(tr,c)
            tc.value=f"=SUM({cl}{hdr_row+1}:{cl}{last})"
            tc.font=Font(name="Arial",bold=True,size=10)
            tc.number_format="#,##0.00"; tc.fill=YELLOW_FILL
            tc.alignment=Alignment(horizontal="right")

    for col_cells in ws.columns:
        w=max((len(str(c.value or "")) for c in col_cells),default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width=min(w+4,32)
    ws.freeze_panes="B6"


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 10 — EXPORT REPORT
# ─────────────────────────────────────────────────────────────────────────────

def export_report(matched, mismatched, miss_gst, miss_pb,
                  pb_agg, gst_agg, all_dups: pd.DataFrame,
                  incomplete_df: pd.DataFrame,
                  out_folder: str, log) -> str:
    out_path = Path(out_folder) / OUTPUT_FILE

    vendor_df = build_vendor_summary(pb_agg, gst_agg) \
                if (not pb_agg.empty or not gst_agg.empty) else pd.DataFrame()
    cross_df  = build_cross_match(miss_gst, miss_pb)
    if not cross_df.empty:
        log(f"  Cross Match: {len(cross_df)} GSTIN(s) found in both missing sheets")

    data_sheets = [
        ("Matched",                  matched,       "Matched"),
        ("Mismatched",               mismatched,    "Mismatched"),
        ("Missing_in_GST_Portal",    miss_gst,      "Missing_in_GST"),
        ("Missing_in_Purchase_Book", miss_pb,       "Missing_in_PB"),
    ]
    if not vendor_df.empty:
        data_sheets.append(("Vendor_Summary",      vendor_df,     "vendor"))
    data_sheets.append(    ("Cross_Match",          cross_df,      "cross"))
    data_sheets.append(    ("Incomplete_Records",   incomplete_df, "incomplete"))

    # Write pandas sheets (skip special ones)
    with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
        for name, df, colour_key in data_sheets:
            if colour_key not in ("vendor", "cross", "incomplete"):
                df.to_excel(writer, sheet_name=name, index=False)

    wb = load_workbook(str(out_path))
    for name, df, colour_key in data_sheets:
        if colour_key == "vendor":
            ws = wb.create_sheet(name)
            _write_vendor_summary_sheet(ws, df, "All Files Combined")
        elif colour_key == "cross":
            ws = wb.create_sheet(name)
            _write_cross_match_sheet(ws, df)
        elif colour_key == "incomplete":
            ws = wb.create_sheet(name)
            _write_incomplete_sheet(ws, df)
        else:
            ws = wb[name]
            _fmt_ws(ws, SHEET_COLOURS.get(colour_key, "595959"))
            if name == "Mismatched":
                _highlight_mismatch(ws, df)
            if ws.max_row > 1:
                _totals_row(ws)

    ws_sum = wb.create_sheet("Summary", 0)
    _write_flat_summary(ws_sum, matched, mismatched, miss_gst, miss_pb,
                        vendor_df, all_dups, incomplete_df)
    wb.save(str(out_path))
    log(f"  Report saved → {out_path}")
    return str(out_path)


def _write_flat_summary(ws, matched, mismatched, miss_gst, miss_pb,
                        vendor_df, all_dups, incomplete_df=None):
    ws["B2"]      = "GSTR-2A vs Purchase Book — Reconciliation Report"
    ws["B2"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws.row_dimensions[2].height = 28
    ws["B3"]      = f"Generated: {datetime.datetime.now().strftime('%d-%m-%Y %H:%M')}"
    ws["B3"].font = Font(name="Arial", italic=True, size=10, color="595959")

    headers = ["Category", "Invoices", "Taxable Value", "IGST", "CGST", "SGST", "ITC At Risk"]
    hf = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    for c, h in enumerate(headers, 2):
        cell = ws.cell(5, c, h)
        cell.font=Font(name="Arial",bold=True,color="FFFFFF",size=10)
        cell.fill=hf; cell.alignment=Alignment(horizontal="center",vertical="center")
        cell.border=THIN

    def _s(df, col):
        use = f"{col}_PB" if f"{col}_PB" in df.columns else col
        return round(df[use].sum() if use in df.columns else 0, 2)
    def _tax(df):
        return round(sum(_s(df, c) for c in ["IGST","CGST","SGST"]), 2)

    rows_data = [
        ("✅ Matched",                matched,    "D9EAD3", False),
        ("🔴 Mismatched (diff > ₹1)", mismatched, "FCE5CD", True),
        ("⚠️ Missing in GST Portal",  miss_gst,   "FFE0E0", True),
        ("⚠️ Missing in Purch. Book", miss_pb,    "EAD1DC", False),
    ]
    for r, (label, df, fhex, at_risk) in enumerate(rows_data, 6):
        fill = PatternFill("solid", start_color=fhex, end_color=fhex)
        vals = [label, len(df), _s(df,"Taxable Value"),
                _s(df,"IGST"), _s(df,"CGST"), _s(df,"SGST"),
                _tax(df) if at_risk else 0.0]
        for c, v in enumerate(vals, 2):
            cell = ws.cell(r, c, v)
            cell.font=Font(name="Arial",size=10,bold=(at_risk and isinstance(v,float) and v>0))
            cell.fill=fill; cell.border=THIN
            cell.alignment=Alignment(
                horizontal="right" if isinstance(v,(int,float)) else "left",
                vertical="center")
            if isinstance(v, float): cell.number_format="#,##0.00"

    total_risk = round(_tax(mismatched) + _tax(miss_gst), 2)
    ws.cell(11, 2).value = f"⚠️  Total ITC At Risk:  ₹{total_risk:,.2f}"
    ws.cell(11, 2).font  = Font(name="Arial", bold=True, size=11, color="C00000")

    # Incomplete records note
    if incomplete_df is not None and not incomplete_df.empty:
        ws.cell(13, 2).value = (f"⚠️  Incomplete Records in Purchase Book: {len(incomplete_df)} row(s) "
                                f"— missing GSTIN or Invoice Number (see 'Incomplete_Records' sheet)")
        ws.cell(13, 2).font  = Font(name="Arial", bold=True, size=10, color="7F6000")

    # Vendor-wise table
    if not vendor_df.empty:
        vr = 16
        ws.cell(vr, 2).value = "Vendor-wise Summary"
        ws.cell(vr, 2).font  = Font(name="Arial", bold=True, size=13, color="1F4E79")
        ws.row_dimensions[vr].height = 22
        ws.cell(vr+1, 2).value = "PB = Purchase Book  |  GST = GST Portal  |  Diff = PB − GST  (red = gap exists)"
        ws.cell(vr+1, 2).font  = Font(name="Arial", italic=True, size=9, color="595959")

        pb_set   = {"TV_PB","IGST_PB","CGST_PB","SGST_PB","Vendor Name (PB)"}
        gst_set  = {"TV_GST","IGST_GST","CGST_GST","SGST_GST","Vendor Name (GST)"}
        diff_set = {"Diff_TV","Diff_IGST","Diff_CGST","Diff_SGST"}
        hfills   = {
            "pb"  : PatternFill("solid",start_color="1F5C2E",end_color="1F5C2E"),
            "gst" : PatternFill("solid",start_color="7030A0",end_color="7030A0"),
            "diff": PatternFill("solid",start_color="C55A11",end_color="C55A11"),
            "key" : PatternFill("solid",start_color="1F4E79",end_color="1F4E79"),
        }
        vcols = list(vendor_df.columns); hdr_r = vr + 2; money_ci = set()
        for c, h in enumerate(vcols, 2):
            cell = ws.cell(hdr_r, c, h)
            cell.font=Font(name="Arial",bold=True,color="FFFFFF",size=10)
            cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=THIN
            if h in pb_set:    cell.fill=hfills["pb"]
            elif h in gst_set: cell.fill=hfills["gst"]
            elif h in diff_set:cell.fill=hfills["diff"]
            else:              cell.fill=hfills["key"]
            if h not in {"GSTIN","Vendor Name (PB)","Vendor Name (GST)"}: money_ci.add(c)

        alt=PatternFill("solid",start_color="EBF3FB",end_color="EBF3FB")
        wht=PatternFill("solid",start_color="FFFFFF",end_color="FFFFFF")
        red=PatternFill("solid",start_color="FFD7D7",end_color="FFD7D7")
        for r,(_, row) in enumerate(vendor_df.iterrows(), hdr_r+1):
            for c,h in enumerate(vcols,2):
                val=row[h]; cell=ws.cell(r,c,val if pd.notna(val) else "")
                cell.border=THIN; cell.font=Font(name="Arial",size=10)
                cell.fill=alt if r%2==0 else wht
                if c in money_ci:
                    cell.number_format="#,##0.00"
                    cell.alignment=Alignment(horizontal="right",vertical="center")
                    if h in diff_set and isinstance(val,(int,float)) and abs(val)>0.01:
                        cell.fill=red; cell.font=Font(name="Arial",size=10,bold=True,color="C00000")
                else:
                    cell.alignment=Alignment(horizontal="left",vertical="center")

        vtr=hdr_r+1+len(vendor_df)
        ws.cell(vtr,2).value="TOTAL"; ws.cell(vtr,2).font=Font(name="Arial",bold=True,size=10)
        ws.cell(vtr,2).fill=YELLOW_FILL
        for c,h in enumerate(vcols,2):
            if c in money_ci:
                cl=get_column_letter(c); tc=ws.cell(vtr,c)
                tc.value=f"=SUM({cl}{hdr_r+1}:{cl}{vtr-1})"
                tc.font=Font(name="Arial",bold=True,size=10)
                tc.number_format="#,##0.00"; tc.fill=YELLOW_FILL
                tc.alignment=Alignment(horizontal="right")

    for c,w in enumerate([2,28,14,28,28,14,14,14,14,14,14,14,14,16,16,16,16],1):
        ws.column_dimensions[get_column_letter(c)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 11 — FULL PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def run_reconciliation(gst_files: list[str], pb_files: list[str],
                       out_folder: str, log, set_progress) -> str:
    set_progress(8)

    # Load all GST files
    log("\n[1] Loading GST Portal files …")
    gst_frames = []
    for fp in gst_files:
        log(f"  Reading: {Path(fp).name}")
        gst_frames.append(load_gst_file(fp, log))
    gst_raw = pd.concat(gst_frames, ignore_index=True) if gst_frames else pd.DataFrame()
    log(f"  GST total: {len(gst_raw):,} rows from {len(gst_files)} file(s)")
    set_progress(25)

    # Load all PB files — now returns (clean, incomplete) tuple
    log("\n[2] Loading Purchase Book files …")
    pb_frames        = []
    incomplete_frames= []
    for fp in pb_files:
        log(f"  Reading: {Path(fp).name}")
        clean_df, inc_df = load_pb_file(fp, log)
        pb_frames.append(clean_df)
        if not inc_df.empty:
            inc_df["Source File"] = Path(fp).name
            incomplete_frames.append(inc_df)

    pb_raw       = pd.concat(pb_frames,         ignore_index=True) if pb_frames         else pd.DataFrame()
    incomplete_df= pd.concat(incomplete_frames, ignore_index=True) if incomplete_frames else pd.DataFrame()
    log(f"  PB total: {len(pb_raw):,} complete rows from {len(pb_files)} file(s)")
    if not incomplete_df.empty:
        log(f"  ⚠️  {len(incomplete_df):,} incomplete rows found across all PB files")
    set_progress(45)

    # Ensure required columns
    for col in KEY_COLS + VALUE_COLS:
        if col not in gst_raw.columns: gst_raw[col] = "" if col in KEY_COLS else 0.0
        if col not in pb_raw.columns:  pb_raw[col]  = "" if col in KEY_COLS else 0.0

    pb_clean  = pb_raw.copy()
    gst_clean = gst_raw.copy()
    set_progress(58)

    log("\n[4] Validating GSTINs …")
    validate_gstins(gst_clean, "GST", log)
    validate_gstins(pb_clean,  "PB",  log)

    log("\n[5] Aggregating invoices …")
    gst_agg = aggregate(gst_clean, "GST", log)
    pb_agg  = aggregate(pb_clean,  "PB",  log)
    set_progress(72)

    log("\n[6] Reconciling …")
    matched, mismatched, miss_gst, miss_pb = reconcile(pb_agg, gst_agg, log)
    set_progress(85)

    log("\n[7] Exporting report …")
    dups_df = pd.DataFrame()
    path = export_report(matched, mismatched, miss_gst, miss_pb,
                         pb_agg, gst_agg, dups_df, incomplete_df, out_folder, log)
    set_progress(100)
    return path


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 12 — GUI
# ─────────────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("GSTR-2A Reconciliation Tool  v2.0")
        self.geometry("820x680")
        self.resizable(False, False)
        self.configure(bg=C["bg"])
        self.gst_files  = []
        self.pb_files   = []
        self.out_folder = tk.StringVar(value=str(Path.home() / "Desktop"))
        self.progress   = tk.DoubleVar(value=0)
        self._build_ui()

    def _build_ui(self):
        hdr = tk.Frame(self, bg=C["primary"], height=64)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text="GSTR-2A Reconciliation Tool",
                 bg=C["primary"], fg=C["white"],
                 font=("Arial",18,"bold")).pack(side="left",padx=24,pady=14)
        tk.Label(hdr, text="v2.0 — Multi-File",
                 bg=C["primary"], fg="#A8C4E0",
                 font=("Arial",10)).pack(side="right",padx=20)

        body = tk.Frame(self, bg=C["bg"])
        body.pack(fill="both", expand=True, padx=20, pady=14)
        left = tk.Frame(body, bg=C["bg"])
        left.pack(side="left", fill="both", expand=True)

        self._multi_file_row(left, "📂  GST Portal Files  (select one or multiple)",
                             "You can select April.xlsx + May.xlsx + June.xlsx at once", "gst")
        self._multi_file_row(left, "📂  Purchase Book Files  (select one or multiple)",
                             "Select all months at once", "pb")

        out_inner = self._card(left, "💾  Output Folder")
        out_row   = tk.Frame(out_inner, bg=C["card"]); out_row.pack(fill="x", pady=4)
        tk.Entry(out_row, textvariable=self.out_folder, font=("Arial",9),
                 relief="solid", highlightthickness=1,
                 highlightbackground=C["border"]).pack(side="left",fill="x",expand=True)
        tk.Button(out_row, text="Browse", command=self._browse_out,
                  bg=C["primary_lt"], fg=C["white"],
                  font=("Arial",9,"bold"), relief="flat",
                  padx=10, cursor="hand2").pack(side="right",padx=(6,0))

        self.run_btn = tk.Button(left, text="▶   Run Reconciliation",
                                 bg=C["primary"], fg=C["white"],
                                 font=("Arial",12,"bold"), relief="flat",
                                 cursor="hand2", pady=10,
                                 activebackground=C["primary_lt"], activeforeground=C["white"],
                                 command=self._start)
        self.run_btn.pack(fill="x", pady=(14,4))

        self.progress_bar = ttk.Progressbar(left, variable=self.progress,
                                            maximum=100, mode="determinate")
        self.progress_bar.pack(fill="x")
        self.status_lbl = tk.Label(left, text="Ready — select files and click Run",
                                   bg=C["bg"], fg=C["subtext"], font=("Arial",9))
        self.status_lbl.pack(anchor="w", pady=(2,0))

        self.badge_frame = tk.Frame(left, bg=C["bg"])
        self.badge_frame.pack(fill="x", pady=(8,0))

        log_card = tk.LabelFrame(body, text=" 📋  Activity Log ",
                                 bg=C["bg"], fg=C["primary"],
                                 font=("Arial",10,"bold"), relief="solid", bd=1)
        log_card.pack(side="right", fill="both", padx=(16,0))
        log_card.configure(width=280); log_card.pack_propagate(False)
        self.log_box = tk.Text(log_card, bg="#0D1117", fg="#58D68D",
                               font=("Consolas",9), relief="flat",
                               wrap="word", state="disabled")
        sb = ttk.Scrollbar(log_card, command=self.log_box.yview)
        self.log_box.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.log_box.pack(fill="both", expand=True, padx=4, pady=4)

    def _multi_file_row(self, parent, title, hint, kind):
        inner = self._card(parent, title)
        tk.Label(inner, text=hint, bg=C["card"], fg=C["subtext"], font=("Arial",8)).pack(anchor="w")
        lb_frame = tk.Frame(inner, bg=C["card"]); lb_frame.pack(fill="x", pady=(4,0))
        lb = tk.Listbox(lb_frame, height=3, font=("Arial",8), relief="solid", bd=1,
                        selectmode="multiple", bg="#F8FAFC", fg=C["text"],
                        highlightthickness=1, highlightbackground=C["border"])
        sb2 = ttk.Scrollbar(lb_frame, orient="vertical", command=lb.yview)
        lb.configure(yscrollcommand=sb2.set)
        sb2.pack(side="right", fill="y"); lb.pack(side="left", fill="x", expand=True)
        btn_row = tk.Frame(inner, bg=C["card"]); btn_row.pack(fill="x", pady=(4,0))

        def browse():
            paths = filedialog.askopenfilenames(
                title=f"Select {'GST Portal' if kind=='gst' else 'Purchase Book'} File(s)",
                filetypes=[("Excel","*.xlsx *.xls"),("All","*.*")])
            if paths:
                if kind == "gst": self.gst_files = list(paths)
                else:             self.pb_files  = list(paths)
                lb.delete(0, "end")
                for p in paths: lb.insert("end", f"  {Path(p).name}")
                self._log(f"Selected {len(paths)} {'GST' if kind=='gst' else 'PB'} file(s):")
                for p in paths: self._log(f"  → {Path(p).name}")

        def clear():
            if kind == "gst": self.gst_files = []
            else:             self.pb_files  = []
            lb.delete(0, "end")

        tk.Button(btn_row, text="Browse Files", command=browse,
                  bg=C["primary_lt"], fg=C["white"], font=("Arial",9,"bold"),
                  relief="flat", padx=10, cursor="hand2").pack(side="left")
        tk.Button(btn_row, text="Clear", command=clear,
                  bg="#E5E7EB", fg=C["text"], font=("Arial",9),
                  relief="flat", padx=8, cursor="hand2").pack(side="left", padx=(6,0))
        count_lbl = tk.Label(btn_row, text="0 files selected",
                             bg=C["card"], fg=C["subtext"], font=("Arial",8))
        count_lbl.pack(side="right")
        self.after(200, lambda: self._watch_count(kind, count_lbl))

    def _watch_count(self, kind, lbl):
        n = len(self.gst_files if kind=="gst" else self.pb_files)
        lbl.configure(text=f"{n} file(s) selected",
                      fg=C["success"] if n>0 else C["subtext"])
        self.after(500, lambda: self._watch_count(kind, lbl))

    def _card(self, parent, title):
        outer = tk.Frame(parent, bg=C["bg"], pady=4); outer.pack(fill="x")
        card  = tk.Frame(outer, bg=C["card"], relief="solid", bd=1,
                         highlightbackground=C["border"], highlightthickness=1)
        card.pack(fill="x")
        tk.Label(card, text=title, bg=C["card"], fg=C["primary"],
                 font=("Arial",10,"bold")).pack(anchor="w",padx=12,pady=(10,4))
        inner = tk.Frame(card, bg=C["card"]); inner.pack(fill="x",padx=12,pady=(0,10))
        return inner

    def _browse_out(self):
        p = filedialog.askdirectory(title="Select Output Folder")
        if p: self.out_folder.set(p); self._log(f"Output folder: {p}")

    def _log(self, msg):
        def _do():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", msg+"\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.after(0, _do)

    def _set_status(self, msg, colour=None):
        self.after(0, lambda: self.status_lbl.configure(text=msg, fg=colour or C["subtext"]))

    def _set_progress(self, val):
        self.after(0, lambda: self.progress.set(val))

    def _start(self):
        if not self.gst_files:
            messagebox.showwarning("Missing Files","Please select at least one GST Portal file."); return
        if not self.pb_files:
            messagebox.showwarning("Missing Files","Please select at least one Purchase Book file."); return
        if not self.out_folder.get():
            messagebox.showwarning("Missing Folder","Please select an output folder."); return

        self.run_btn.configure(state="disabled", text="⏳  Running …")
        self.log_box.configure(state="normal"); self.log_box.delete("1.0","end")
        self.log_box.configure(state="disabled")
        self.progress.set(0)
        for w in self.badge_frame.winfo_children(): w.destroy()

        threading.Thread(target=self._pipeline,
                         args=(list(self.gst_files), list(self.pb_files), self.out_folder.get()),
                         daemon=True).start()

    def _pipeline(self, gst_files, pb_files, out_folder):
        try:
            self._log("━━━ Starting Reconciliation ━━━")
            self._log(f"GST files : {len(gst_files)}")
            self._log(f"PB files  : {len(pb_files)}")
            self._set_status("Running …", C["primary"])

            path = run_reconciliation(gst_files, pb_files, out_folder,
                                      self._log, self._set_progress)

            wb = load_workbook(path)
            total_m = max(wb["Matched"].max_row - 2, 0)                  if "Matched"                  in wb.sheetnames else 0
            total_x = max(wb["Mismatched"].max_row - 2, 0)               if "Mismatched"               in wb.sheetnames else 0
            total_g = max(wb["Missing_in_GST_Portal"].max_row - 2, 0)    if "Missing_in_GST_Portal"    in wb.sheetnames else 0
            total_p = max(wb["Missing_in_Purchase_Book"].max_row - 2, 0) if "Missing_in_Purchase_Book" in wb.sheetnames else 0
            total_i = max(wb["Incomplete_Records"].max_row - 5, 0)       if "Incomplete_Records"       in wb.sheetnames else 0

            self._set_status("✅  Report saved: "+path, C["success"])
            self._set_progress(100)

            def _do():
                for w in self.badge_frame.winfo_children(): w.destroy()
                data = [
                    (f"✅ {total_m:,}\nMatched",    C["success"], C["green_bg"]),
                    (f"🔴 {total_x:,}\nMismatched", C["warning"], C["orange_bg"]),
                    (f"⚠️ {total_g:,}\nMiss. GST",  C["danger"],  C["red_bg"]),
                    (f"⚠️ {total_p:,}\nMiss. PB",   C["purple"],  C["purple_bg"]),
                    (f"⚠️ {total_i:,}\nIncomplete", "#7F6000",    "#FFF9E6"),
                ]
                for text,fg,bg in data:
                    tk.Label(self.badge_frame, text=text, bg=bg, fg=fg,
                             font=("Arial",9,"bold"), relief="solid", bd=1,
                             width=10, justify="center", padx=4, pady=6
                             ).pack(side="left",padx=4,expand=True,fill="x")
            self.after(0, _do)

            self.after(0, lambda: messagebox.showinfo(
                "Done ✅",
                f"Reconciliation complete!\n\n"
                f"File saved: {path}\n\n"
                f"  ✅ Matched               : {total_m:,}\n"
                f"  🔴 Mismatched            : {total_x:,}\n"
                f"  ⚠️  Missing in GST Portal : {total_g:,}\n"
                f"  ⚠️  Missing in Purch. Book: {total_p:,}\n"
                f"  ⚠️  Incomplete Records    : {total_i:,}"
            ))

        except Exception as e:
            self._log(f"\n❌ ERROR: {e}")
            self._set_status(f"❌ Error: {e}", C["danger"])
            self.after(0, lambda: messagebox.showerror("Error", str(e)))
        finally:
            self.after(0, lambda: self.run_btn.configure(
                state="normal", text="▶   Run Reconciliation"))


if __name__ == "__main__":
    App().mainloop()